#!/usr/bin/env python

import os
import re
import sys
import getopt
import readline
import string
import pyexcel as p
import xlrd

from openpyxl import load_workbook


################################
class file_group:
   def __init__(self, _index,_target_dir,_pre_dir_path,_new_file_path):
      self.dir_index = _index
      self.target_dir= _target_dir
      self.pre_dir   = _pre_dir_path
      self.new_file  = _new_file_path
      print( self.dir_index,self.target_dir,self.pre_dir,self.new_file )

################################

def merge_files( _pre_file,_cur_file,_out_file ):
    "合并文件指定内容，输出到第三个文件"
    pre_wb = xlrd.open_workbook( _pre_file )
    pre_wb_ws = pre_wb.sheet_by_index(0)

    cur_wb = xlrd.open_workbook( _cur_file )
    cur_wb_ws = cur_wb.sheet_by_name("2-社会与经济信息表（4）")

    target_wb = xlrd.open_workbook( _out_file )
    target_wb_ws = target_wb.sheet_by_name("完整版")

    rows = 30
    cols = 4
    for r in range(4,rows):
        pre_val  = pre_wb_ws.cell_value( r ,4) 
        cur_val  = cur_wb_ws.cell_value( r ,4)
        # target_wb_ws.cell( r ,4).value = pre_val
        # target_wb_ws.cell( r ,5).value = cur_val
        print('%s' % r, pre_val,cur_val )
    print("**************")
    new_out_file = "n_"+_out_file
    target_wb.save( new_out_file )


# def merge_files( _pre_file,_cur_file,_out_file ):
#     "合并文件指定内容，输出到第三个文件"
#     pre_wb = load_workbook( _pre_file )
#     pre_wb_ws = pre_wb.active

#     cur_wb = load_workbook( _cur_file )
#     cur_wb_ws = cur_wb["2-社会与经济信息表（4）"]

#     target_wb = load_workbook( _out_file )
#     target_wb_ws = target_wb["完整版"]
#     target_wb_ws.unmerge_cells('A1:E1')
#     target_wb_ws.unmerge_cells('A2:D2')
#     target_wb_ws.unmerge_cells('A3:C3')

#     rows = 30
#     cols = 4
#     for r in range(4,rows):
#         pre_val  = pre_wb_ws.cell( r ,4).value 
#         cur_val  = cur_wb_ws.cell( r ,4).value 
#         target_wb_ws.cell( r ,4).value = pre_val
#         target_wb_ws.cell( r ,5).value = cur_val
#         print('%s' % r, pre_val,cur_val )
#     print("**************")
#     target_wb_ws.merge_cells('A1:E1')
#     target_wb_ws.merge_cells('A2:D2')
#     target_wb_ws.merge_cells('A3:C3')

#     new_out_file = "n_"+_out_file
#     target_wb.save( new_out_file )

#################################

pairs_list = []


if __name__ =='__main__' : 
    pre_dir_list = os.listdir( './原数据' )
    print("**************")
    
    new_list = os.listdir( './新数据' )
    print("**************")

    target_dir_list = os.listdir( './最终表格' )
    for target_dir in target_dir_list :
        # 遍历所有的目标文件夹，获取index和文件夹路径
        target_dir_str_list = re.findall(r"[0-9]+", target_dir )
        target_index = target_dir_str_list[0]
        target_dir = './最终表格/' + target_dir
        # print( '%s' % target_dir,target_index )
        new_file = ''
        pre_dir  = ''

        # 以index为准，逐一获取新文件，和老文件夹的路径
        for new_file_name in new_list:
            new_file_name_str_list = re.findall(r"[0-9]+", new_file_name )
            new_file_name_index = new_file_name_str_list[0]
            if ( new_file_name_index == target_index ):
                new_file = './新数据/'+new_file_name
                # print( '%s' % target_dir ,new_file_name ) 

        for pre_dir_name in pre_dir_list:
            pre_dir_name_str_list = re.findall(r"[0-9]+", pre_dir_name )
            pre_dir_name_index = pre_dir_name_str_list[0]
            if ( pre_dir_name_index == target_index ):
                pre_dir = './原数据/'+pre_dir_name

        # 从目标文件夹中获取最终的标的文件tt_f
        file_in_target_dir_list = os.listdir( target_dir )
        # print(file_in_target_dir_list)
        for ff in file_in_target_dir_list:
            # 获取第一个目标文件
            if '社会与经济信息表' in ff :
                tt_f = target_dir + '/' + ff
        
        # 获取原数据中的标的文件 pp_f
        file_in_pre_dir_list = os.listdir( pre_dir )
        print(file_in_pre_dir_list)
        for pp in file_in_pre_dir_list:
            # 获取第一个目标文件
            if '社会与经济信息表' in pp :
                pp_f = pre_dir + '/' + pp

        print( '>>> %s' % pp_f,new_file,tt_f)
        # 处理三个文件
        merge_files(pp_f,new_file,tt_f)

    





